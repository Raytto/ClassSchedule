#!/usr/bin/env python
# coding: utf-8

# In[ ]:





# # 1. 数据设置

# ## 1.1 时空信息

# In[1]:


# 班级序号
all_class_indexes = [i for i in range(1, 11)]
# all_class_indexes = [1]
# 天序号
all_days_indexes = [i for i in range(1, 8)]
# 每天天内的课程序号
all_course_indexes = [i for i in range(0, 14)]


# ## 1.2 时间化分

# In[2]:


all_section = ["dawn", "morning", "afternoon", "evening"]
indexes_by_section = {
    "dawn": [0],
    "morning": [1, 2, 3, 4, 5],
    "afternoon": [6, 7, 8, 9],
    "evening": [10, 11, 12, 13],
}
max_section_len = max([len(x) for x in indexes_by_section.values()])
# time用(day,index)表示
time_by_section = {
    "dawn": [(the_day, 0) for the_day in range(1, 7)],  # 星期1到6都有早自习
    "morning": [
        (the_day, the_index)
        for the_day in range(1, 6)
        for the_index in indexes_by_section["morning"]
    ],
    "afternoon": [
        (the_day, the_index)
        for the_day in range(1, 6)
        for the_index in indexes_by_section["afternoon"]
    ],
    "evening": [
        (the_day, the_index)
        for the_day in [1, 2, 3, 4, 5, 7]
        for the_index in indexes_by_section["evening"]
    ],  # 周天有晚自习
}


# In[3]:


# time_by_section["evening"]
max_section_len


# ## 1.3 课程信息

# In[4]:


# 课程类型
all_subjects = [
    "语文",
    "美术",
    "政治",
    "地理",
    "英语",
    "心理",
    "生物",
    "体育",
    "化学",
    "音乐",
    "历史",
    "数学",
    "物理",
    "劳动",
    "班会",
    "",  # 无课
]


# In[5]:


# 一些不需要老师的课程
no_teacher_subjects = ["劳动", ""]


# ## 1.4 教师信息

# In[6]:


# 教师信息：名字：负责的科目和科目对应的班级
teachers_duties = {
    "teacher_28": {"语文": [1, 3]},
    "teacher_59": {"数学": [1, 7]},
    "teacher_18": {"英语": [1, 9]},
    "teacher_15": {"政治": [1, 3]},
    "teacher_6": {"历史": [1, 4, 6]},
    "teacher_46": {"地理": [1, 2, 6], "班会": [1]},
    "teacher_50": {"美术": [1, 2, 8, 9, 10]},
    "teacher_51": {"化学": [1, 2]},
    "teacher_13": {"物理": [1, 2, 4]},
    "teacher_39": {"生物": [1, 9, 10]},
    "teacher_29": {"心理": [1, 2, 8, 9, 10]},
    "teacher_52": {"音乐": [1, 2, 9, 10]},
    "teacher_22": {"体育": [1, 2, 8, 9, 10]},
    "teacher_24": {"英语": [2, 6]},
    "teacher_53": {"数学": [2]},
    "teacher_2": {"语文": [2], "班会": [2]},
    "teacher_35": {"生物": [2, 4, 8]},
    "teacher_57": {"历史": [2, 3, 7]},
    "teacher_37": {"政治": [2, 7]},
    "teacher_26": {"英语": [3, 7]},
    "teacher_32": {"数学": [3], "班会": [3]},
    "teacher_54": {"地理": [3, 4, 5]},
    "teacher_20": {"物理": [3]},
    "teacher_1": {"音乐": [3, 4, 5, 6, 7, 8]},
    "teacher_23": {"生物": [3, 6], "班会": [6]},
    "teacher_42": {"化学": [3, 9]},
    "teacher_43": {"美术": [3, 4, 5, 6, 7]},
    "teacher_12": {"体育": [3, 4, 5, 6, 7]},
    "teacher_44": {"心理": [3, 4, 5, 6, 7]},
    "teacher_7": {"语文": [4, 8]},
    "teacher_8": {"数学": [4, 10]},
    "teacher_16": {"英语": [4, 10]},
    "teacher_55": {"化学": [4, 7], "班会": [4]},
    "teacher_33": {"政治": [4, 5, 10]},
    "teacher_34": {"语文": [5, 9]},
    "teacher_19": {"数学": [5, 6]},
    "teacher_40": {"物理": [5, 7], "班会": [7]},
    "teacher_41": {"英语": [5, 8]},
    "teacher_45": {"生物": [5, 7]},
    "teacher_25": {"化学": [5, 6], "班会": [5]},
    "teacher_30": {"历史": [5, 8, 9, 10]},
    "teacher_9": {"语文": [6]},
    "teacher_27": {"物理": [6, 9]},
    "teacher_17": {"政治": [6]},
    "teacher_21": {"语文": [7, 10]},
    "teacher_10": {"地理": [7, 8, 9, 10]},
    "teacher_47": {"数学": [8], "班会": [8]},
    "teacher_14": {"政治": [8, 9]},
    "teacher_36": {"物理": [8, 10], "班会": [10]},
    "teacher_31": {"化学": [8, 10]},
    "teacher_38": {"数学": [9], "班会": [9]},
}


# ## 1.5 班级信息

# In[7]:


# 班级分了物理和历史两个方向
all_foucs = ["physics", "history"]
# 方向和班级序号的关系
class_foucs_info = {"physics": [3, 4, 5, 6, 7, 8, 9, 10], "history": [1, 2]}


# ## 1.6 各班的科目数量要求
# 针对 "对某种方向的班级，某时间集合内，要求某科目数量刚好为n" 这种需求

# In[8]:


# 用一个数据结构描述,每一个要求为(foucs,[(day1,index1),(day2,index2)], aim_subject, aim_num)
subject_num_requirements = []


# ### 1.6.1 早自习要求
# 每班 周一到周六 ，早自习课程要求 语文英语各3节

# In[9]:


morning_time_set = [(the_day, 0) for the_day in range(1, 7)]
for the_foucs in all_foucs:
    for the_subject in ["语文", "英语"]:
        subject_num_requirements.append((the_foucs, morning_time_set, the_subject, 3))


# ### 1.6.2 正课数量要求

# In[10]:


# 正课程数量要求
subject_requirement_in_regular = {
    "physics": {
        "语文": 6,
        "数学": 7,
        "英语": 8,
        "物理": 5,
        "化学": 4,
        "生物": 3,
        "政治": 2,
        "历史": 2,
        "地理": 2,
        "体育": 2,
        "心理": 1,
        "劳动": 1,
        "美术": 1,
        "音乐": 1,
    },
    "history": {
        "语文": 6,
        "数学": 7,
        "英语": 8,
        "物理": 2,
        "化学": 2,
        "生物": 2,
        "政治": 4,
        "历史": 4,
        "地理": 4,
        "体育": 2,
        "心理": 1,
        "劳动": 1,
        "美术": 1,
        "音乐": 1,
    },
}

for the_foucs in subject_requirement_in_regular.keys():
    for the_subject, the_num in subject_requirement_in_regular[the_foucs].items():
        subject_num_requirements.append(
            (
                the_foucs,
                (time_by_section["morning"] + time_by_section["afternoon"]),
                the_subject,
                the_num,
            )
        )


# ### 1.6.3 晚自习数量要求

# In[11]:


### 晚自数要求
subject_num_in_evening = {
    "physics": {
        "语文": 3,
        "数学": 4,
        "英语": 3,
        "物理": 4,
        "化学": 3,
        "生物": 3,
        "政治": 1,
        "历史": 1,
        "地理": 1,
        "班会": 1,
    },
    "history": {
        "语文": 3,
        "数学": 4,
        "英语": 3,
        "物理": 1,
        "化学": 1,
        "生物": 1,
        "政治": 3,
        "历史": 4,
        "地理": 3,
        "班会": 1,
    },
}
for the_foucs in subject_num_in_evening.keys():
    for the_subject, the_num in subject_num_in_evening[the_foucs].items():
        subject_num_requirements.append(
            (the_foucs, time_by_section["evening"], the_subject, the_num)
        )


# In[12]:


# time_by_section["evening"]


# ### 1.6.4 休息时间要求

# In[13]:


rest_time_set = []
for the_index in (
    indexes_by_section["morning"]
    + indexes_by_section["afternoon"]
    + indexes_by_section["evening"]
):
    rest_time_set.append((6, the_index))
for the_index in (
    indexes_by_section["dawn"]
    + indexes_by_section["morning"]
    + indexes_by_section["afternoon"]
):
    rest_time_set.append((7, the_index))
for the_foucs in all_foucs:
    subject_num_requirements.append((the_foucs, rest_time_set, "", len(rest_time_set)))


# In[14]:


# rest_time_set


# ### 1.6.5 特殊课程要求

# In[15]:


for the_foucs in all_foucs:
    # 星期天第一节晚自习上班会
    subject_num_requirements.append((the_foucs, [(7, 10)], "班会", 1))
    # 星期一下午最后一节上劳动
    subject_num_requirements.append((the_foucs, [(1, 9)], "劳动", 1))


# In[16]:


# subject_num_requirements


# In[17]:


len(rest_time_set)


# ## 1.2 特殊信息

# ### 1.2.1 教研时间段

# In[18]:


jiaoyan_day_section = {
    "语文": (4, "afternoon"),
    "数学": (3, "morning"),
    "英语": (1, "afternoon"),
    "物理": (4, "afternoon"),
    "化学": (2, "morning"),
    "生物": (2, "morning"),
    "政治": (5, "morning"),
    "历史": (5, "morning"),
    "地理": (4, "afternoon"),
    "体育": (1, "morning"),
    "音乐": (1, "morning"),
    "美术": (1, "morning"),
    "心理": (1, "morning"),
}


# # 2. 预处理

# In[19]:


# 生成一个通过班级和课程找到老师的映射表，方便后续查询
# 需要要求一个班的同一科目只有一个老师负责
map_class_subject_to_teacher = {}  # by (the_class_index,the_subject)

for the_teacher, duties in teachers_duties.items():
    for the_subject, the_classes in duties.items():
        for the_class_index in the_classes:
            map_class_subject_to_teacher[(the_class_index, the_subject)] = the_teacher

# 一些课程不需要老师
for the_class_index in all_class_indexes:
    for the_subject in no_teacher_subjects:
        map_class_subject_to_teacher[(the_class_index, the_subject)] = ""


# In[20]:


# str(map_class_subject_to_teacher)


# In[21]:


# 生成一些方便后续使用的变量
all_teachers = teachers_duties.keys()
classes_of_teacher = {}
subjects_of_teacher = {}
for the_teacher, duties in teachers_duties.items():
    classes_of_teacher[the_teacher] = []
    subjects_of_teacher[the_teacher] = []
    for the_subject, the_classes in duties.items():
        classes_of_teacher[the_teacher] = list(
            set(classes_of_teacher[the_teacher]) | set(the_classes)
        )
        if the_subject not in subjects_of_teacher[the_teacher]:
            subjects_of_teacher[the_teacher].append(the_subject)


# In[22]:


# classes_of_teacher


# # 3. 设置模型
# 口语课是单周口语，双周英语，单独用一个变量来记

# ## 3.1 求解目标变量

# In[23]:


from ortools.sat.python import cp_model

# 创建模型
model = cp_model.CpModel()

# 创建目标变量：每班每节课的科目
# model.course_vars = Var(
#     all_class_indexes,
#     all_days_indexes,
#     all_course_indexes,
#     all_subjects,
#     domain=Binary,
# )
course_vars = {}
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            for the_subject in all_subjects:
                course_vars[the_class,the_day,the_index,the_subject] = model.NewBoolVar(f'course_vars{the_class}_{the_day}_{the_index}_{the_subject}')
                
# 创建目标变量：每班每节课是否是外教课
# model.is_spoken_course = Var(
#     model.all_class_indexes,
#     model.all_days_indexes,
#     model.all_course_indexes,
#     domain=Binary,
# )
is_spoken_course = {}
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            is_spoken_course[the_class,the_day,the_index] = model.NewBoolVar(f'is_spoken_course{the_class}_{the_day}_{the_index}')


# ## 3.1 预设一些通用约束函数

# ### 3.1.1 逻辑与约束

# In[24]:


def logic_cons_and(model,a,xs):
    model.AddBoolAnd(xs).OnlyEnforceIf(a)
    model.AddBoolOr([x.Not() for x in xs]).OnlyEnforceIf(a.Not())


# ## 3.2 设置辅助变量

# In[25]:


## 老师是否在上课或者教研的变量
teacher_time_vars = {}
for the_teacher in all_teachers:
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            teacher_time_vars[the_teacher,the_day,the_index] = model.NewBoolVar(f'teacher_time_vars{the_teacher}_{the_day}_{the_index}')

## bigM，用于一些约束中使非线性约束转化为线性约束
bigM = 10000

# 可以有连续课程的section
indexes_could_consecutive = (
    indexes_by_section["morning"][:-1]
    + indexes_by_section["afternoon"][:-1]
    + indexes_by_section["evening"][:-1]
)


# In[26]:


# 创建变量指针哪些课程将连续上(对班级而言)
will_consecutive_by_class_time_subject = {}
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_index in indexes_could_consecutive:
            for the_subject in all_subjects:
                will_consecutive_by_class_time_subject[the_class,the_day,the_index,the_subject] = model.NewBoolVar(f'will_consecutive_by_class_time_subject_{the_class}_{the_day}_{the_index}_{the_subject}')


for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_index in indexes_could_consecutive:
            for the_subject in all_subjects:
                logic_cons_and(
                    model,
                    will_consecutive_by_class_time_subject[
                        the_class, the_day, the_index, the_subject
                    ],
                    [
                        course_vars[the_class, the_day, the_index, the_subject],
                        course_vars[
                            the_class, the_day, the_index + 1, the_subject
                        ],
                    ],
                )

# 创建变量计算每班每个day_section内每科有多少连续次数(对班级而言)
sections_could_consecutive = [x for x in all_section if x != "dawn"]
consecutive_num_by_class_day_section_subject = {}
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_section in sections_could_consecutive:
            for the_subject in all_subjects:
                consecutive_num_by_class_day_section_subject[the_class,the_day,the_section,the_subject] = model.NewIntVar(0, max_section_len, f'consecutive_num_by_class_day_section_subject_{the_class}_{the_day}_{the_section}_{the_subject}')


for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_section in sections_could_consecutive:
            for the_subject in all_subjects:
                model.Add(
                    consecutive_num_by_class_day_section_subject[
                        the_class, the_day, the_section, the_subject
                    ]
                    == sum(
                        will_consecutive_by_class_time_subject[
                            the_class, the_day, the_index, the_subject
                        ]
                        for the_index in indexes_by_section[the_section][:-1]
                    )
                )

# 创建变量计算每班每个day_section内每科有多少节课(对班级而言)
num_by_class_day_section_subject = {}
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_section in all_section:
            for the_subject in all_subjects:
                num_by_class_day_section_subject[the_class,the_day,the_section,the_subject] = model.NewIntVar(0, max_section_len, f'num_by_class_day_section_subject_{the_class}_{the_day}_{the_section}_{the_subject}')


# 创建变量对应每班每个day_section内每科是否有课(对班级而言)
any_by_class_day_section_subject = {}
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_section in all_section:
            for the_subject in all_subjects:
                any_by_class_day_section_subject[the_class,the_day,the_section,the_subject] = model.NewBoolVar(f'any_by_class_day_section_subject_{the_class}_{the_day}_{the_section}_{the_subject}')

for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_section in all_section:
            for the_subject in all_subjects:
                model.Add(
                    num_by_class_day_section_subject[
                        the_class, the_day, the_section, the_subject
                    ]
                    == sum(
                        course_vars[the_class, the_day, the_index, the_subject]
                        for the_index in indexes_by_section[the_section]
                    )
                )
                # 用  b<=M*a
                #    b>=0.5-M(1-a)
                # 约束a=1 if b >0 else a=0
                
                a = any_by_class_day_section_subject[
                    the_class, the_day, the_section, the_subject
                ]
                b = num_by_class_day_section_subject[
                    the_class, the_day, the_section, the_subject
                ]
                model.Add(b>0).OnlyEnforceIf(a)
                model.Add(b==0).OnlyEnforceIf(a.Not())
                # model.Add(b <= a * bigM)
                # model.Add(b >= 1 - bigM * (1 - a))


# In[27]:


# 创建变量计算每老师每个day_section内每科有多少节课(对老师而言)
num_by_teacher_day_section = {}
for the_teacher in all_teachers:
    for the_day in all_days_indexes:
        for the_section in all_section:
            num_by_teacher_day_section[the_teacher,the_day,the_section] = model.NewIntVar(0, max_section_len, f'num_by_teacher_day_section_{the_teacher}_{the_day}_{the_section}')

# 创建变量对应每老师每个day_section内每科是否有课(对老师而言)
any_by_teacher_day_section = {}
for the_teacher in all_teachers:
    for the_day in all_days_indexes:
        for the_section in all_section:
            any_by_teacher_day_section[the_teacher,the_day,the_section] = model.NewBoolVar( f'any_by_teacher_day_section{the_teacher}_{the_day}_{the_section}')

for the_teacher in all_teachers:
    for the_day in all_days_indexes:
        for the_section in all_section:
            # for the_subject in all_subjects:
            has_jiaoyan = 0
            for the_subject in set(teachers_duties[the_teacher].keys()) & set(
                jiaoyan_day_section.keys()
            ):
                if (
                    the_section == jiaoyan_day_section[the_subject][1]
                    and the_day == jiaoyan_day_section[the_subject][0]
                ):
                    has_jiaoyan = 1

            model.Add(
                num_by_teacher_day_section[the_teacher, the_day, the_section]
                == sum(
                    course_vars[the_class, the_day, the_index, the_subject]
                    for the_index in indexes_by_section[the_section]
                    for the_subject, the_class_list in teachers_duties[
                        the_teacher
                    ].items()
                    for the_class in the_class_list
                )
                + has_jiaoyan
            )
            # 用  b<=M*a
            #    b>=0.5-M(1-a)
            # 约束a=1 if b >0 else a=0
            a = any_by_teacher_day_section[the_teacher, the_day, the_section]
            b = num_by_teacher_day_section[the_teacher, the_day, the_section]
            model.Add(b>0).OnlyEnforceIf(a)
            model.Add(b==0).OnlyEnforceIf(a.Not())
            # model.add(b <= a * bigM)
            # model.add(b >= 1 - bigM * (1 - a))


# # 4. 加约束

# ## 4.1 最基本约束

# ### 4.1.1 每教师同时至多上一节课

# In[ ]:





# In[28]:


# 定义约束
for the_teacher in all_teachers:
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            has_jiaoyan = 0
            for the_subject in set(teachers_duties[the_teacher].keys()) & set(
                jiaoyan_day_section.keys()
            ):
                if (
                    the_index in indexes_by_section[jiaoyan_day_section[the_subject][1]]
                    and the_day == jiaoyan_day_section[the_subject][0]
                ):
                    has_jiaoyan = 1
                    break
            model.Add(
                teacher_time_vars[the_teacher, the_day, the_index] 
                == (
                    sum(
                        course_vars[(the_class, the_day, the_index, the_subject)]
                        for the_subject, the_class_list in teachers_duties[the_teacher].items()
                        for the_class in set(the_class_list)&set(all_class_indexes)
                        if the_class in all_class_indexes
                    )
                    + has_jiaoyan
                )
            )
            # if the_teacher == "teacher_24" and the_day ==1 and the_index == 7:
            #     print(has_jiaoyan)


# In[29]:


# model.teacher_time_constraint.pprint()


# In[30]:


# # teachers_duties
# model.addCons(course_vars[(2, 1, 6, "英语")] == 1)
# model.addCons(course_vars[(2, 1, 7, "英语")] == 1)
# model.addCons(course_vars[(2, 1, 8, "英语")] == 1)
# model.addCons(
#     teacher_time_vars[(the_teacher, the_day, the_index)]
#     == quicksum([course_vars[(2, 1, 6, "英语")] , course_vars[(6, 1, 6, "英语")]])
#     + (1 == jiaoyan_day_section["英语"][0]) * (1 in [1, 2, 3])
# )


# In[31]:


# (1 == jiaoyan_day_section["英语"][0]) * (1 in [1, 2, 3])


# ### 4.1.2 每班同时仅上一节课
# 包含空课

# In[32]:


# 定义约束

for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            model.Add(
                sum(
                    course_vars[the_class, the_day, the_index, the_subject]
                    for the_subject in all_subjects
                )
                == 1
            )


# ## 4.2 各时段课程数量约束

# In[33]:


for the_class in all_class_indexes:
    the_class_foucs = "history"
    if the_class_index in class_foucs_info["physics"]:
        the_class_foucs = "physics"
    for the_foucs, time_set, aim_subject, aim_num in subject_num_requirements:
        if the_foucs == the_class_foucs:
            # print(the_foucs,"  ",time_set,"  ",aim_subject,"  ",aim_num)
            model.add(
                sum(
                    course_vars[
                        (
                            the_class,
                            aim_day,
                            aim_index,
                            aim_subject,
                        )
                    ]
                    for aim_day, aim_index in time_set
                )
                == aim_num
            )


# In[ ]:





# ## 4.3 特殊约束

# ### 4.3.1 口语课要求

# In[34]:


# 约束 : 对每个班的任意时间而言，单周的口语在双周对应英语
for a_class in all_class_indexes:
    for a_day in all_days_indexes:
        for a_course_Indices in all_course_indexes:
            model.Add(
                course_vars[(a_class, a_day, a_course_Indices, "英语")]
                >= is_spoken_course[(a_class, a_day, a_course_Indices)]
            )

# 约束 : 每个班每周需要一节口语课
for the_class in all_class_indexes:
    model.Add(
        sum(
            is_spoken_course[(the_class, the_day, the_index)]
            for the_day in all_days_indexes
            for the_index in all_course_indexes
        )
        == 1
    )

# 约束 : 同一时间只能有一节口语课
for the_day in all_days_indexes:
    for the_index in all_course_indexes:
        model.Add(
            sum(
                is_spoken_course[(the_class, the_day, the_index)]
                for the_class in all_class_indexes
            )
            <= 1
        )

# 约束 : 口语仅在周一周五正课阶段上
for the_class in all_class_indexes:
    for the_day in [2, 3, 4, 6, 7]:
        for the_index in all_course_indexes:
            model.Add(
                is_spoken_course[(the_class, the_day, the_index)] == 0
            )
    for the_day in [1, 5]:
        for the_index in indexes_by_section["dawn"] + indexes_by_section["evening"]:
            model.Add(
                is_spoken_course[(the_class, the_day, the_index)] == 0
            )


# ### 4.3.2 每班每科正课每天数量限制
# 语数外 至少1节，至多2节
# 其他科目至多1节

# In[35]:


main_subjects = ["语文","数学","英语"]
for the_class in all_class_indexes:
    for the_day in [1,2,3,4,5]:
        for the_subject in filter(lambda x: x != "", all_subjects):
            if the_subject in main_subjects:
                model.Add(
                    sum(
                        course_vars[
                            (
                                the_class,
                                the_day,
                                the_index,
                                the_subject,
                            )
                        ]
                        for the_index in (
                            indexes_by_section["morning"] + indexes_by_section["afternoon"]
                        )
                    )
                    <= 2
                )
                model.Add(
                    sum(
                        course_vars[
                            (
                                the_class,
                                the_day,
                                the_index,
                                the_subject,
                            )
                        ]
                        for the_index in (
                            indexes_by_section["morning"] + indexes_by_section["afternoon"]
                        )
                    )
                    >= 1
                )
            else:
                model.Add(
                    sum(
                        course_vars[
                            (
                                the_class,
                                the_day,
                                the_index,
                                the_subject,
                            )
                        ]
                        for the_index in (
                            indexes_by_section["morning"] + indexes_by_section["afternoon"]
                        )
                    )
                    <= 1
                )


# ### 4.3.3 每班每科晚自习每天至多两节

# In[36]:


for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_subject in filter(lambda x: x != "", all_subjects):
            model.Add(
                sum(
                    course_vars[
                        (
                            the_class,
                            the_day,
                            the_index,
                            the_subject,
                        )
                    ]
                    for the_index in indexes_by_section["evening"]
                )
                <= 2
            )


# ### 4.3.4 每个班而言每个时间段内的相同两节课需要连续

# In[37]:


# 创建连续约束，每班每个day_section内，同课程要么没有，要么数量是连续次数+1
# if k then a=b-1 转线性约束
# a>=b-1-M*(1-k)
# a<=b-1+M*(1-k)
for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_section in sections_could_consecutive:
            for the_subject in all_subjects:
                a = consecutive_num_by_class_day_section_subject[
                    the_class, the_day, the_section, the_subject
                ]
                b = num_by_class_day_section_subject[
                    the_class, the_day, the_section, the_subject
                ]
                k = any_by_class_day_section_subject[
                    the_class, the_day, the_section, the_subject
                ]
                model.Add(a==b-1).OnlyEnforceIf(k)
                # model.Add(a >= b - 1 - bigM * (1 - k))
                # model.Add(a <= b - 1 + bigM * (1 - k))


# ### 4.3.5 每个班每天每科不会上下午都上

# In[38]:


for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_subject in filter(lambda x: x != "", all_subjects):
            model.Add(
                any_by_class_day_section_subject[
                    the_class, the_day, "morning", the_subject
                ]
                + any_by_class_day_section_subject[
                    the_class, the_day, "afternoon", the_subject
                ]
                <= 1
            )


# ### 4.3.6 周五下午行政不排课

# In[39]:


### 周五下午行政不排课
xingzheng_teachers = [
    "teacher_56",
    "teacher_3",
    "teacher_48",
    "teacher_53",
    "teacher_49",
    "teacher_4",
    "teacher_29",
    "teacher_20",
    "teacher_11",
    "teacher_58",
    "teacher_5",
]
for the_teacher in xingzheng_teachers:
    if the_teacher not in teachers_duties.keys():
        # print(f" teacher {the_teacher} has no class")
        continue
    for the_subject, the_classes in teachers_duties[the_teacher].items():
        for the_class in the_classes:
            for the_index in indexes_by_section["afternoon"]:
                model.Add(
                    course_vars[(the_class, 5, the_index, the_subject)] == 0
                )


# ### 4.3.7 教研时间不上课

# In[40]:


# for the_subject, all_the_times in jaoyan_requirement.items():
#     for the_day, the_section in all_the_times:
#         for the_index in index_section[the_section]:
#             for the_teacher in all_teachers:
#                 if teachers_and_subjects_map[the_teacher] == the_subject:
#                     for the_class in teachers_and_class_map[the_teacher]:
#                         model.Add(
#                             course_vars[(the_class, the_day, the_index, the_subject)]
#                             == 0
#                         )


# ### 4.3.8 如果下午是教研会晚上一般不安排晚自习

# In[41]:


for the_subject, day_section in jiaoyan_day_section.items():
    the_day, the_section = day_section
    if the_section == "afternoon":
        for the_index in indexes_by_section["evening"]:
            for the_class in all_class_indexes:
                model.Add(
                    course_vars[(the_class, the_day, the_index, the_subject)] == 0
                )


# ### 4.3.9 各班各科每周的晚自习前后均匀
# 前两节晚自习数量 与 后两节晚自习数量的差值低于1

# In[42]:


half_evening = {"early": [10, 11], "late": [12, 13]}

half_evening_num_by_class_half_subject = {}
for half_name in half_evening.keys():
    for the_class in all_class_indexes:
        for the_subject in all_subjects:
            half_evening_num_by_class_half_subject[half_name,the_class,the_subject] = model.NewIntVar(0, max_section_len* len(all_days_indexes), f'half_evening_num_by_class_half_subject{half_name}_{the_class}_{the_subject}')

for the_class in all_class_indexes:
    for the_subject in filter(lambda x: x != "", all_subjects):
        for the_half_name, the_half_indexes in half_evening.items():
            model.Add(
                half_evening_num_by_class_half_subject[
                    the_half_name, the_class, the_subject
                ]
                == sum(
                    course_vars[the_class, the_day, the_index, the_subject]
                    for the_day in all_days_indexes
                    for the_index in the_half_indexes
                )
            )
        model.Add(
            half_evening_num_by_class_half_subject[
                "early", the_class, the_subject
            ]
            - half_evening_num_by_class_half_subject[
                "late", the_class, the_subject
            ]
            <= 1
        )
        model.Add(
            half_evening_num_by_class_half_subject["late", the_class, the_subject]
            - half_evening_num_by_class_half_subject[
                "early", the_class, the_subject
            ]
            <= 1
        )


# In[ ]:





# # 5. 加目标

# ## 5.1 子目标

# ### 5.1.1 教师需要出现的section总数尽量少

# In[43]:


teacher_num_by_day_section = {}
for the_day in all_days_indexes:
    for the_section in all_section:
        teacher_num_by_day_section[the_day, the_section] = model.NewIntVar(0, len(all_teachers), f'teacher_num_by_day_section_{the_day}_{the_section}')

for the_day in all_days_indexes:
    for the_section in all_section:
        model.Add(
            teacher_num_by_day_section[the_day, the_section]
            == sum(
                any_by_teacher_day_section[the_teacher, the_day, the_section]
                for the_teacher in all_teachers
            )
        )
total_teacher_shows_day_section = model.NewIntVar(
    0, len(all_teachers) * len(all_days_indexes) * len(all_section),f"total_teacher_shows_day_section"
)
model.Add(
    total_teacher_shows_day_section
    == sum(
        teacher_num_by_day_section[the_day, the_section]
        for the_day in all_days_indexes
        for the_section in all_section
    )
)


# ### 5.1.2 老师每天课程集中得尽量多
# 靠目标函数中的系数体现我们希望各个出现的倾向性

# In[44]:


care_section_from_to = [
    ("dawn","morning"),
    ("dawn","afternoon"),
    ("dawn","evening"),
    ("morning","afternoon"),
    ("afternoon","evening"),
    ("morning","evening"),
]
teacher_times_by_section_from_to = {}
teacher_day_section_from_to = {}
for section_from,section_to in care_section_from_to:
    teacher_times_by_section_from_to[section_from,section_to] = model.NewIntVar(
        0, len(all_teachers)*len(all_days_indexes), 
        f'teacher_times_by_section_from_to_{section_from}_{section_to}'
    )

for section_from,section_to in care_section_from_to:
    for the_teacher in all_teachers:
        for the_day in [1,2,3,4,5]:
            teacher_day_section_from_to[the_teacher,the_day,section_from,section_to] = model.NewBoolVar(
                f'teacher_day_section_from_to_{the_teacher}_{the_day}_{section_from}_{section_to}'
            )

for the_teacher in all_teachers:
    for the_day in [1,2,3,4,5]:
        a = teacher_day_section_from_to[the_teacher, the_day, "dawn","morning"]
        b = any_by_teacher_day_section[the_teacher, the_day, "dawn"]
        c = any_by_teacher_day_section[the_teacher, the_day, "morning"]
        d = any_by_teacher_day_section[the_teacher, the_day, "afternoon"].Not()
        e = any_by_teacher_day_section[the_teacher, the_day, "evening"].Not()
        model.AddBoolAnd([b, c,d,e]).OnlyEnforceIf(a)   # b and c 必须为 True 时，a 为 True
        model.AddBoolOr([b.Not(),c.Not(),d.Not(),e.Not()]).OnlyEnforceIf(a.Not())  

model.Add(
    teacher_times_by_section_from_to["dawn","morning"]
    == sum (
        teacher_day_section_from_to[the_teacher, the_day, "dawn","morning"]
        for the_teacher in all_teachers
        for the_day in [1,2,3,4,5]
    )
)

for the_teacher in all_teachers:
    for the_day in [1,2,3,4,5]:
        a = teacher_day_section_from_to[the_teacher, the_day, "dawn","afternoon"]
        b = any_by_teacher_day_section[the_teacher, the_day, "dawn"]
        d = any_by_teacher_day_section[the_teacher, the_day, "afternoon"]
        e = any_by_teacher_day_section[the_teacher, the_day, "evening"].Not()
        model.AddBoolAnd([b,d,e]).OnlyEnforceIf(a)   # b and c 必须为 True 时，a 为 True
        model.AddBoolOr([b.Not(),d.Not(),e.Not()]).OnlyEnforceIf(a.Not())  

model.Add(
    teacher_times_by_section_from_to["dawn","afternoon"]
    == sum (
        teacher_day_section_from_to[the_teacher, the_day, "dawn","afternoon"]
        for the_teacher in all_teachers
        for the_day in [1,2,3,4,5]
    )
)

for the_teacher in all_teachers:
    for the_day in [1,2,3,4,5]:
        a = teacher_day_section_from_to[the_teacher, the_day, "dawn","evening"]
        e = any_by_teacher_day_section[the_teacher, the_day, "evening"]
        model.AddBoolAnd([b,e]).OnlyEnforceIf(a)   # b and c 必须为 True 时，a 为 True
        model.AddBoolOr([b.Not(),e.Not()]).OnlyEnforceIf(a.Not())  

model.Add(
    teacher_times_by_section_from_to["dawn","evening"]
    == sum (
        teacher_day_section_from_to[the_teacher, the_day, "dawn","evening"]
        for the_teacher in all_teachers
        for the_day in [1,2,3,4,5]
    )
)
        
for the_teacher in all_teachers:
    for the_day in [1,2,3,4,5]:
        a = teacher_day_section_from_to[the_teacher, the_day, "morning","afternoon"]
        b = any_by_teacher_day_section[the_teacher, the_day, "dawn"].Not()
        f = any_by_teacher_day_section[the_teacher, the_day, "morning"]
        d = any_by_teacher_day_section[the_teacher, the_day, "afternoon"]
        e = any_by_teacher_day_section[the_teacher, the_day, "evening"].Not()
        model.AddBoolAnd([b,d,e,f]).OnlyEnforceIf(a)   # b and c 必须为 True 时，a 为 True
        model.AddBoolOr([b.Not(),d.Not(),e.Not(),f.Not()]).OnlyEnforceIf(a.Not())  

model.Add(
    teacher_times_by_section_from_to["morning","afternoon"]
    == sum (
        teacher_day_section_from_to[the_teacher, the_day, "morning","afternoon"]
        for the_teacher in all_teachers
        for the_day in [1,2,3,4,5]
    )
)

for the_teacher in all_teachers:
    for the_day in [1,2,3,4,5]:
        a = teacher_day_section_from_to[the_teacher, the_day, "morning","evening"]
        b = any_by_teacher_day_section[the_teacher, the_day, "dawn"].Not()
        f = any_by_teacher_day_section[the_teacher, the_day, "morning"]
        e = any_by_teacher_day_section[the_teacher, the_day, "evening"]
        model.AddBoolAnd([b,e,f]).OnlyEnforceIf(a)   # b and c 必须为 True 时，a 为 True
        model.AddBoolOr([b.Not(),e.Not(),f.Not()]).OnlyEnforceIf(a.Not())  

model.Add(
    teacher_times_by_section_from_to["morning","evening"]
    == sum (
        teacher_day_section_from_to[the_teacher, the_day, "morning","evening"]
        for the_teacher in all_teachers
        for the_day in [1,2,3,4,5]
    )
)

for the_teacher in all_teachers:
    for the_day in [1,2,3,4,5]:
        a = teacher_day_section_from_to[the_teacher, the_day, "afternoon","evening"]
        b = any_by_teacher_day_section[the_teacher, the_day, "dawn"].Not()
        f = any_by_teacher_day_section[the_teacher, the_day, "morning"].Not()
        c = any_by_teacher_day_section[the_teacher, the_day, "afternoon"]
        e = any_by_teacher_day_section[the_teacher, the_day, "evening"]
        model.AddBoolAnd([b,e,c,f]).OnlyEnforceIf(a)   # b and c 必须为 True 时，a 为 True
        model.AddBoolOr([b.Not(),e.Not(),c.Not(),f.Not()]).OnlyEnforceIf(a.Not())  

model.Add(
    teacher_times_by_section_from_to["afternoon","evening"]
    == sum (
        teacher_day_section_from_to[the_teacher, the_day, "afternoon","evening"]
        for the_teacher in all_teachers
        for the_day in [1,2,3,4,5]
    )
)


# ## 5.2 综合目标函数

# In[45]:


model.Minimize(
    teacher_num_by_day_section[1, "evening"] * 100
    + teacher_num_by_day_section[2, "evening"] * 100
    + teacher_num_by_day_section[3, "evening"] * 100
    + teacher_num_by_day_section[4, "evening"] * 100
    + teacher_num_by_day_section[5, "evening"] * 100
    + teacher_num_by_day_section[7, "evening"] * 1000
    + total_teacher_shows_day_section * 10
    - teacher_times_by_section_from_to["dawn","morning"] * 5
    - teacher_times_by_section_from_to["afternoon","evening"] *5
    + teacher_times_by_section_from_to["dawn","afternoon"] * 5
    + teacher_times_by_section_from_to["morning","afternoon"] * 5
    + teacher_times_by_section_from_to["morning","evening"] * 20
    + teacher_times_by_section_from_to["dawn","evening"] * 200
)


# # 6. 求解

# In[46]:


import json
load_course_var = {}
with open('course_save_var_31499.json', 'r') as f:
    load_course_var = json.load(f)

for the_class in all_class_indexes:
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            for the_subject in all_subjects:
                str_key = str((the_class,the_day,the_index,the_subject))
                model.AddHint(course_vars[the_class,the_day,the_index,the_subject],load_course_var[str_key])


# ## 6.1 使用OR-TOOLS求解

# In[47]:


class SolutionPrinter(cp_model.CpSolverSolutionCallback):
    def __init__(self):
        cp_model.CpSolverSolutionCallback.__init__(self)
        print(f"Start to find Solution")
        self.__solution_count = 0

    def OnSolutionCallback(self):
        # This method is called each time the solver finds a solution.
        print(f"Solution {self.__solution_count} : Objective Value = {self.ObjectiveValue()} Time = {self.WallTime()} s")
        # Access variables using the Value method, e.g., `self.Value(var_name)`
        self.__solution_count += 1


# In[48]:


solution_printer = SolutionPrinter()
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = 18000
status = solver.Solve(model,solution_printer)


# In[49]:


if status == cp_model.OPTIMAL:
    print("方案 OPTIMAL")
elif status == cp_model.FEASIBLE:
    print("方案 FEASIBLE")
elif status == cp_model.INFEASIBLE:
    print("INFASIBLE")
elif status == cp_model.MODEL_INVALID:
    print("MODEL_INVALID")
elif status == cp_model.UNKNOWN:
    print("UNKNOWN")


# In[50]:


# 保存结果

save_course_var = {}
if status in [cp_model.OPTIMAL,cp_model.FEASIBLE]:
    for the_key in course_vars.keys():
        str_key = str(the_key)
        save_course_var[str_key] = solver.Value(course_vars[the_key])

    # 将临时解写入文件
    with open(f'course_save_var_{int(solver.ObjectiveValue())}.json', 'w') as f:
        json.dump(save_course_var, f)


# In[51]:


# teacher_time_vars[the_teacher, the_day, the_index] 


# ## 6.2 抽查解信息

# In[52]:


print(solver.Value(teacher_time_vars[("teacher_24", 1, 6)]))
print(solver.Value(course_vars[(2, 1, 7, "英语")]))
print(map_class_subject_to_teacher[2, "英语"])
# print(model.any_by_class_day_section_subject[1, 2, "morning", "数学"].value)


# In[53]:


# 班级信息x
the_class = 2
for the_day in all_days_indexes:
    for the_index in all_course_indexes:
        for the_subject in all_subjects:
            if model.course_vars[(the_class, the_day, the_index, the_subject)].value == 1:
                print(
                    f"the_class {the_class} the_day {the_day} the_index {the_index} the_subject {the_subject}"
                )


# # 7. 输出到 xlsx

# ## 7.1 基本设置

# In[54]:


# 输出到 xlsx
from copy import copy
import numpy as np
import pandas as pd
from openpyxl import load_workbook


# In[55]:


def copy_sheet(source_sheet, target_sheet):
    # 复制单元格内容和格式
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            target_cell.value = cell.value
            if cell.has_style:
                target_cell._style = copy(cell._style)
            if cell.hyperlink:
                target_cell.hyperlink = copy(cell.hyperlink)
            if cell.comment:
                target_cell.comment = copy(cell.comment)

    # 复制行高
    for row, dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row].height = dimension.height

    # 复制列宽
    for col, dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col].width = dimension.width

    # 指定所有列的宽度
    fixed_width = 20  # 设置的宽度值

    # 获取工作表的最大列数
    max_column = target_sheet.max_column

    # 遍历所有列并设置宽度
    for col in range(1, max_column + 1):
        col_letter = target_sheet.cell(row=1, column=col).column_letter  # 获取列字母
        target_sheet.column_dimensions[col_letter].width = fixed_width

    # 复制合并单元格
    if source_sheet.merged_cells:
        target_sheet.merged_cells = source_sheet.merged_cells


# ## 7.2 输出班级课表

# In[56]:


# 输出班级表
workbook_class = load_workbook("sample_class.xlsx")
map_from_index_to_row = [3, 5, 7, 8, 10, 11, 15, 16, 17, 18, 21, 22, 23, 24]
map_from_day_to_col = [1, 2, 3, 4, 5, 6, 7, 8]
working_teachers = []
sample_sheet = workbook_class["排版"]
for the_class in all_class_indexes:
    the_sheet = workbook_class.create_sheet(title=f"高2024级{the_class}班")
    copy_sheet(sample_sheet, the_sheet)
    the_sheet.cell(row=1, column=1).value = f"高2024级{the_class}班 课表"
    for the_day in all_days_indexes:
        for the_index in all_course_indexes:
            for the_subject in all_subjects:
                has_course = solver.Value(course_vars[
                    (the_class, the_day, the_index, the_subject)
                ])
                if has_course:
                    is_spoken = solver.Value(is_spoken_course[
                        (the_class, the_day, the_index)
                    ])
                    if is_spoken == 1:
                        content = "口语(单周)\n英语(双周)"
                    else:
                        content = the_subject
                    teacher = map_class_subject_to_teacher[(the_class, the_subject)]
                    # print(f'row {row_index} col {col_index} content {content}')
                    #                     print(f'row {the_index}')
                    the_sheet.cell(
                        row=map_from_index_to_row[the_index],
                        column=map_from_day_to_col[the_day],
                    ).value = f"{content}\n{teacher}"
                    if teacher not in working_teachers:
                        working_teachers.append(teacher)
workbook_class.remove(sample_sheet)
workbook_class.save(f"result_class_{int(solver.ObjectiveValue())}.xlsx")


# ## 7.3 输出老师课表

# In[57]:


# 输出老师表
workbook_teacher = load_workbook("sample_class.xlsx")
map_from_index_to_row = [3, 5, 7, 8, 10, 11, 15, 16, 17, 18, 21, 22, 23, 24]
map_from_day_to_col = [1, 2, 3, 4, 5, 6, 7, 8]
sample_sheet = workbook_teacher["排版"]
for the_teacher in working_teachers:
    if the_teacher != "":
        # print(the_teacher)
        the_sheet = workbook_teacher.create_sheet(title=f"{the_teacher}")
        copy_sheet(sample_sheet, the_sheet)
        the_sheet.cell(row=1, column=1).value = f"{the_teacher} 课表"
        for the_day in all_days_indexes:
            for the_index in all_course_indexes:
                # 教研
                for the_subject, the_times in teachers_duties[the_teacher].items():
                    the_jiaoyan_day_section = jiaoyan_day_section.get(the_subject, None)
                    if (
                        the_jiaoyan_day_section != None
                        and the_day == the_jiaoyan_day_section[0]
                        and the_index in indexes_by_section[the_jiaoyan_day_section[1]]
                    ):
                        the_sheet.cell(
                            row=map_from_index_to_row[the_index],
                            column=map_from_day_to_col[the_day],
                        ).value = f"{the_subject}教研\n{the_teacher}"
                # 其他课
                for the_class in all_class_indexes:
                    for the_subject in all_subjects:
                        has_course = solver.Value(course_vars[
                            (the_class, the_day, the_index, the_subject)
                        ])
                        if (
                            has_course == 1
                            and map_class_subject_to_teacher[(the_class, the_subject)]
                            == the_teacher
                        ):
                            is_spoken = solver.Value(is_spoken_course[
                                (the_class, the_day, the_index)
                            ])
                            if is_spoken == 1:
                                content = "口语(单周)\n英语(双周)"
                            else:
                                content = the_subject
                            the_sheet.cell(
                                row=map_from_index_to_row[the_index],
                                column=map_from_day_to_col[the_day],
                            ).value = f"{content}\n{the_class}班\n{the_teacher}"

# 单独输出口语sheet
the_sheet = workbook_teacher.create_sheet(title="口语")
copy_sheet(sample_sheet, the_sheet)
for the_day in all_days_indexes:
    for the_index in all_course_indexes:
        for the_class in all_class_indexes:
            is_spoken = solver.Value(is_spoken_course[(the_class, the_day, the_index)])
            if is_spoken == 1:
                content = "口语(单周)"
                the_sheet.cell(
                    row=map_from_index_to_row[the_index],
                    column=map_from_day_to_col[the_day],
                ).value = f"{content}\n{the_class}班"
workbook_teacher.remove(sample_sheet)
workbook_teacher.save(f"result_teacher_{int(solver.ObjectiveValue())}.xlsx")


# In[58]:


# jiaoyan_day_section = {
#     "语文": (4, "afternoon"),
#     "数学": (3, "morning"),
#     "英语": (1, "afternoon"),
#     "物理": (4, "afternoon"),
#     "化学": (2, "morning"),
#     "生物": (2, "morning"),
#     "政治": (5, "morning"),
#     "历史": (5, "morning"),
#     "地理": (4, "afternoon"),
#     "体育": (1, "morning"),
#     "音乐": (1, "morning"),
#     "美术": (1, "morning"),
#     "心理": (1, "morning"),
# }

